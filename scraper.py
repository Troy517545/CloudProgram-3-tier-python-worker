import requests
#from selenium import webdriver
from bs4 import BeautifulSoup
#from selenium.webdriver.common.action_chains import ActionChains
#from selenium.common.exceptions import NoSuchElementException
#from selenium.webdriver.common.keys import Keys
import itertools
import re,os, sys
from openpyxl import load_workbook
from xlutils.copy import copy

CURRENT_ROW = 0

loc = sys.argv[1]

d_list=[]

rb = load_workbook(loc)
ws = rb.active
column = ws['E']

for x in range(CURRENT_ROW,301):
    #print(column[x].value)
    d_list.append(column[x].value)

    #driver = webdriver.Chrome(executable_path='/opt/python/current/app/chromedriver')


    cnt = 0
    for company in d_list:
        if cnt == 1: break
        cnt += 1
        #    var = "_"
        #    path = var.join(company.split(" "))
        #    path = "reports/"+path
        #    if not os.path.isdir(path):
        #        os.makedirs(path)


        str_ = "+"
        seq = str_.join(company.split(" "))

        web_page = requests.get('https://beta.companieshouse.gov.uk/search?q='+seq)
        soup = BeautifulSoup(web_page.text, 'html.parser')
        company_url = soup.find("ul",id="results").find("li",recursive=False).find("h3").find("a")['href']
        FS_page = requests.get('https://beta.companieshouse.gov.uk/'+company_url+'/filing-history')
        if (not FS_page): break

        print('https://beta.companieshouse.gov.uk'+company_url+'/filing-history')
        #   driver.get('https://beta.companieshouse.gov.uk'+company_url+'/filing-history')


        #        try:
        #           #To find whether the checkbox element is selected or not
        #          isChecked = driver.find_element_by_id('filter-category-accounts').is_selected()
        #     except:
        #        print('No Account Data :O')
        #       cell = ws.cell(row= CURRENT_ROW+1 , column = 10)
        #      cell.value = "Nan"
        #     rb.save(loc)
        #    break

        #if isChecked is False:
        #   driver.find_element_by_id("filter-category-accounts").click()
        #  driver.find_element_by_id('filing-history-tab').click()
        #To find whether the checkbox element is selected or not
        # isChecked = driver.find_element_by_id('filter-category-accounts').is_selected()
        # if isChecked is False:
        #     driver.find_element_by_id("filter-category-accounts").click()

        #  FS_page = driver.page_source
        soup = BeautifulSoup(FS_page.text, 'html.parser')
        try:
            fsTable = soup.find('table', id='fhTable').find_all('tr')
        except:
            print('Report Not Found :(')
            cell = ws.cell(row= CURRENT_ROW+1 , column = 10)
            cell.value = "None"
            rb.save(loc)
            break


            for idx, tr in enumerate(fsTable):
                if(idx == 0): continue
                if(idx == 6): break
                td = tr.contents
                date = td[1].string

                pdfTag = td[7].find('a', class_='download')

                if(pdfTag): pdfURL = pdfTag['href']

                pdf = requests.get('https://beta.companieshouse.gov.uk' + pdfURL)
                file_name = '_'.join(company.split(' ')) + '_' + ''.join(re.split(' |\n', date))
                print("download %s to RDS" % file_name)
                # f = open(path+'/'+file_name+'.pdf', 'wb+')
                # for chunk in pdf.iter_content(chunk_size=32):
                #     f.write(chunk)



                #        try:
                #            driver.find_element_by_id('nextButton').click()
                #            print('Next Page...')
                #        except NoSuchElementException:
                #            print('No Next Page~')
                #            cell = ws.cell(row= CURRENT_ROW+1 , column = 10)
                #            cell.value = "Done"
                #            rb.save(loc)
                #            break

                CURRENT_ROW += 1